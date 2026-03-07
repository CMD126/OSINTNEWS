"""
Reporter module - terminal output + HTML / CSV / JSON / Excel export.
Supports optional AI analysis section in HTML (from RAG pipeline).
"""

from __future__ import annotations
import os
import csv
import json
import re
import html as html_lib
from datetime import datetime
from typing import Optional

# ANSI colors for CLI
_CYAN   = "\033[96m"
_GREEN  = "\033[92m"
_YELLOW = "\033[93m"
_BLUE   = "\033[94m"
_WHITE  = "\033[97m"
_GREY   = "\033[90m"
_RED    = "\033[91m"
_RESET  = "\033[0m"


# ---------------------------------------------------------------------------
# Terminal output
# ---------------------------------------------------------------------------

def print_results(results: list, target: str) -> None:
    if not results:
        print(f"\n{_RED}[!] No results found.{_RESET}")
        return

    print(f"\n{_GREEN}{'=' * 70}{_RESET}")
    print(f"{_CYAN}  Results for: {_WHITE}{target}{_CYAN}  |  Total: {_WHITE}{len(results)}{_RESET}")
    print(f"{_GREEN}{'=' * 70}{_RESET}")

    current_cat = None
    idx = 1
    for r in results:
        cat = r.get("category", "Unknown")
        if cat != current_cat:
            current_cat = cat
            print(f"\n{_YELLOW}  [{cat}]{_RESET}")
            print(f"  {_GREY}{'─' * 60}{_RESET}")

        title   = r.get("title", "No title")
        url     = r.get("href",  "")
        body    = r.get("body",  "") or ""
        snippet = (body[:150] + "...") if len(body) > 150 else body

        print(f"\n  {_WHITE}{idx}. {title}{_RESET}")
        print(f"     {_BLUE}{url}{_RESET}")
        if snippet:
            print(f"     {_GREY}{snippet}{_RESET}")
        idx += 1


# ---------------------------------------------------------------------------
# HTML report  (with optional AI analysis block)
# ---------------------------------------------------------------------------

def save_report(
    results:    list,
    target:     str,
    output_dir: str = "reports",
    analysis=None,          # Optional[AIAnalysis]
) -> str:
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe      = "".join(c if c.isalnum() or c in "-_" else "_" for c in target)
    path      = os.path.join(output_dir, f"osintnews_{safe}_{timestamp}.html")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(_build_html(results, target, timestamp, analysis))
    return path


def save_report_to(
    results:  list,
    target:   str,
    path:     str,
    analysis=None,          # Optional[AIAnalysis]
) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(_build_html(results, target, timestamp, analysis))
    return path


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

def save_csv(results: list, path: str) -> str:
    fields = ["target", "category", "title", "href", "body", "query"]
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)
    return path


# ---------------------------------------------------------------------------
# JSON export
# ---------------------------------------------------------------------------

def save_json(results: list, path: str) -> str:
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(results, fh, indent=2, ensure_ascii=False)
    return path


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def save_excel(results: list, path: str) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OSINTNEWS Results"

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(bold=True, color="58A6FF")

    headers = ["#", "Target", "Category", "Title", "URL", "Snippet", "Dork Query"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(results, 1):
        ws.append([
            i,
            r.get("target",   ""),
            r.get("category", ""),
            r.get("title",    ""),
            r.get("href",     ""),
            (r.get("body", "") or "")[:300],
            r.get("query",    ""),
        ])

    widths = [5, 20, 25, 50, 60, 80, 60]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Markdown → HTML (simple, no external library)
# ---------------------------------------------------------------------------

def _md_to_html(text: str) -> str:
    """Pure function: converts the LLM's markdown output to safe HTML."""
    # Escape HTML entities first
    text = html_lib.escape(text)

    # Headings
    text = re.sub(r'^## (.+)$',  r'<h2>\1</h2>', text, flags=re.MULTILINE)
    text = re.sub(r'^### (.+)$', r'<h3>\1</h3>', text, flags=re.MULTILINE)

    # Bold
    text = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', text)

    # Bullet points
    text = re.sub(r'^[-•] (.+)$', r'<li>\1</li>', text, flags=re.MULTILINE)

    # Wrap consecutive <li> blocks in <ul>
    text = re.sub(r'((?:<li>.*?</li>\n?)+)', r'<ul>\1</ul>', text, flags=re.DOTALL)

    # Citation badges  [Source N]
    text = re.sub(
        r'\[Source\s+(\d+)\]',
        r'<span class="cite">[Source \1]</span>',
        text,
    )

    # Warning markers
    text = text.replace('⚠ Insufficient data.', '<span class="warn">⚠ Insufficient data.</span>')

    # Paragraphs (double newline → <p>)
    blocks = text.split('\n\n')
    parts  = []
    for block in blocks:
        block = block.strip()
        if block and not block.startswith('<h') and not block.startswith('<ul') \
                 and not block.startswith('<li'):
            block = f'<p>{block}</p>'
        parts.append(block)
    return '\n'.join(parts)


# ---------------------------------------------------------------------------
# HTML builder
# ---------------------------------------------------------------------------

def _build_html(results: list, target: str, timestamp: str, analysis=None) -> str:
    cats: dict = {}
    for r in results:
        cat = r.get("category", "Uncategorized")
        cats.setdefault(cat, []).append(r)

    # --- AI analysis block ---
    ai_block = ""
    if analysis and analysis.succeeded:
        risk_color = analysis.risk_color_hex
        ai_html    = _md_to_html(analysis.summary)
        cited_list = "".join(
            f'<li><a href="{html_lib.escape(u)}" target="_blank" rel="noopener">{html_lib.escape(u)}</a></li>'
            for u in analysis.cited_sources
        )
        ai_block = f"""
        <section class="ai-section">
            <div class="ai-header">
                <span class="ai-label">&#129302; AI Intelligence Report</span>
                <span class="ai-model">{html_lib.escape(analysis.model)}</span>
                <span class="risk-badge" style="background:{risk_color}20;color:{risk_color};border-color:{risk_color}">
                    {analysis.risk_level}
                </span>
            </div>
            <div class="ai-body">
                {ai_html}
            </div>
            {"<div class='cited-sources'><strong>Cited Sources:</strong><ul>" + cited_list + "</ul></div>" if cited_list else ""}
        </section>
        """
    elif analysis and not analysis.succeeded:
        ai_block = f"""
        <section class="ai-section ai-error">
            <div class="ai-header">
                <span class="ai-label">&#129302; AI Analysis — Error</span>
            </div>
            <p class="error-msg">{html_lib.escape(analysis.error or "Unknown error")}</p>
        </section>
        """

    # --- News result cards ---
    cards = ""
    for cat, items in cats.items():
        cards += f"""
        <section class="cat">
            <h2 class="cat-title">{html_lib.escape(cat)}
                <span class="badge">{len(items)}</span>
            </h2>
            <div class="cards">
        """
        for item in items:
            t   = html_lib.escape(item.get("title",  "No title"))
            url = html_lib.escape(item.get("href",   "#"))
            b   = html_lib.escape((item.get("body",  "") or "")[:300])
            q   = html_lib.escape(item.get("query",  ""))
            tgt = html_lib.escape(item.get("target", ""))
            cards += f"""
                <article class="card">
                    <a class="card-title" href="{url}" target="_blank" rel="noopener">{t}</a>
                    <div class="card-meta"><span class="card-target">&#128269; {tgt}</span></div>
                    <div class="card-url">{url}</div>
                    <div class="card-body">{b}</div>
                    <span class="card-query">{q}</span>
                </article>
            """
        cards += "</div></section>"

    safe_target = html_lib.escape(target)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>OSINTNEWS Plus &mdash; {safe_target}</title>
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#0d1117;color:#c9d1d9;font-family:'Segoe UI',system-ui,sans-serif;padding:24px;line-height:1.6}}
a{{color:#58a6ff}}
a:hover{{color:#79c0ff}}

header{{border-bottom:1px solid #30363d;padding-bottom:20px;margin-bottom:32px}}
h1{{font-size:1.9rem;color:#58a6ff}}
.meta{{color:#8b949e;font-size:.88rem;margin-top:6px}}
.meta strong{{color:#e6edf3}}
.stats{{display:flex;gap:16px;margin-top:16px;flex-wrap:wrap}}
.stat{{background:#161b22;border:1px solid #30363d;border-radius:8px;padding:10px 20px}}
.stat-num{{font-size:1.6rem;color:#3fb950;font-weight:700}}
.stat-label{{font-size:.75rem;color:#8b949e}}

/* AI Section */
.ai-section{{background:#161b22;border:1px solid #30363d;border-radius:12px;padding:24px;margin-bottom:40px}}
.ai-section.ai-error{{border-color:#f85149}}
.ai-header{{display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap}}
.ai-label{{font-size:1.1rem;font-weight:700;color:#c9d1d9}}
.ai-model{{font-size:.75rem;color:#8b949e;background:#21262d;padding:3px 8px;border-radius:4px;font-family:Consolas,monospace}}
.risk-badge{{font-size:.8rem;font-weight:700;padding:4px 12px;border-radius:20px;border:1px solid;letter-spacing:.5px}}
.ai-body h2{{color:#f78166;font-size:1.05rem;margin:20px 0 8px}}
.ai-body h3{{color:#d29922;font-size:.95rem;margin:14px 0 6px}}
.ai-body p{{color:#c9d1d9;margin-bottom:10px;line-height:1.7}}
.ai-body ul{{margin:8px 0 12px 20px}}
.ai-body li{{color:#c9d1d9;margin-bottom:4px;line-height:1.6}}
.ai-body strong{{color:#f0f6fc}}
.cite{{background:#1f6feb30;color:#79c0ff;font-size:.8rem;padding:1px 5px;border-radius:3px;font-family:Consolas,monospace}}
.warn{{color:#f78166;font-style:italic}}
.error-msg{{color:#f85149;font-family:Consolas,monospace;font-size:.9rem}}
.cited-sources{{margin-top:16px;padding-top:16px;border-top:1px solid #30363d}}
.cited-sources strong{{color:#8b949e;font-size:.85rem}}
.cited-sources ul{{margin:8px 0 0 16px}}
.cited-sources li{{font-size:.78rem;margin-bottom:3px;color:#8b949e}}

/* News Cards */
.cat{{margin-bottom:44px}}
.cat-title{{font-size:1.1rem;color:#f78166;margin-bottom:14px;display:flex;align-items:center;gap:8px}}
.badge{{background:#21262d;color:#8b949e;font-size:.72rem;padding:2px 8px;border-radius:20px;font-weight:400}}
.cards{{display:grid;gap:12px}}
.card{{background:#161b22;border:1px solid #30363d;border-radius:10px;padding:16px;transition:border-color .18s}}
.card:hover{{border-color:#58a6ff}}
.card-title{{display:block;color:#58a6ff;font-size:.97rem;font-weight:600;text-decoration:none;margin-bottom:5px}}
.card-title:hover{{color:#79c0ff;text-decoration:underline}}
.card-meta{{font-size:.72rem;color:#8b949e;margin-bottom:4px}}
.card-target{{color:#d29922}}
.card-url{{color:#3fb950;font-size:.72rem;margin-bottom:8px;word-break:break-all}}
.card-body{{color:#8b949e;font-size:.86rem;margin-bottom:8px}}
.card-query{{display:block;background:#1c2128;border:1px solid #30363d;color:#6e7681;font-size:.68rem;font-family:Consolas,monospace;padding:4px 8px;border-radius:4px;word-break:break-all}}

footer{{border-top:1px solid #30363d;padding-top:20px;margin-top:40px;color:#484f58;font-size:.78rem;text-align:center}}
</style>
</head>
<body>
<header>
  <h1>&#128269; OSINTNEWS Plus</h1>
  <div class="meta">Target(s): <strong>{safe_target}</strong> &bull; Generated: {timestamp}</div>
  <div class="stats">
    <div class="stat"><div class="stat-num">{len(results)}</div><div class="stat-label">Results</div></div>
    <div class="stat"><div class="stat-num">{len(cats)}</div><div class="stat-label">Categories</div></div>
    {"<div class='stat'><div class='stat-num' style='color:" + analysis.risk_color_hex + "'>" + analysis.risk_level + "</div><div class='stat-label'>Risk Level</div></div>" if analysis and analysis.succeeded else ""}
  </div>
</header>

{ai_block}

{cards}

<footer>Generated by OSINTNEWS Plus &mdash; RAG Intelligence Engine</footer>
</body>
</html>"""
